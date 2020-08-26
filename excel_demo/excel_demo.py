from openpyxl import Workbook

from openpyxl import load_workbook

from openpyxl.writer.excel import ExcelWriter

from openpyxl.utils import get_column_letter

import datetime
from random import choice
from time import time


def read():
    wb = load_workbook('D:/文档/项目/贵阳TOCC/dwd数据字典.xlsx')
    # 显示所有表名
    # print(wb.sheetnames)

    # 遍历所有表
    for sheet in wb:
        print('标题：' + sheet.title)
        cols = sheet['1']
        for col in cols:
            print('第一行值：' + col.value)

        for row in sheet['A']:
            print('A列值：' + row.value)

        for row in sheet['A2':'A3']:
            print('A2-A3：' + row[0].value)
        break

    # 获得最大列和最大行
    sheet = wb.active
    print('最大行数：' + sheet.max_row)
    print('最大列数：' + sheet.max_column)


def write():
    # 实例化
    wb = Workbook()
    # 激活 worksheet
    sheet = wb.active
    sheet['A1'] = 'A1'

    # 合并单元格， 往左上角写入数据即可
    # sheet.merge_cells('B1:G1')  # 合并一行中的几个单元格
    # sheet.merge_cells('A1:C3')  # 合并一个矩形区域中的单元格

    # 拆分单元格
    # sheet.unmerge_cells('A1:C3')

    wb.save('./test.xlsx')


def demo():
    wb = Workbook()
    ws = wb.active

    addr = './demo.xlsx'
    # wb = load_workbook()
    # ws = wb.create_sheet()

    # 第一行输入
    ws.append(['TIME', 'TITLE', 'A-Z'])

    # 输入内容（500行数据）
    for i in range(500):
        TIME = datetime.datetime.now().strftime("%H:%M:%S")
        TITLE = str(time())
        A_Z = get_column_letter(choice(range(1, 50)))
        ws.append([TIME, TITLE, A_Z])

    # 获取最大行
    row_max = ws.max_row
    # 获取最大列
    con_max = ws.max_column
    # 把上面写入内容打印在控制台
    for j in ws.rows:  # we.rows 获取每一行数据
        for n in j:
            print(n.value, end="\t")  # n.value 获取单元格的值
        print()
    # 保存，save（必须要写文件名（绝对地址）默认 py 同级目录下，只支持 xlsx 格式）
    wb.save(addr)


if __name__ == '__main__':
    # read()
    # write()
    demo()
